# -*- coding: utf-8 -*-
# @Time    : 2019/12/4 17:35
# @Author  : man.jiang
from openpyxl import load_workbook
import openpyxl
class Write_excel(object):

    '''修改excel数据'''
    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active  # 激活sheet

    # def write(self, row_n, col_n, value):
    #  '''写入数据，如(2,3，"hello"),第二行第三列写入数据"hello"'''
    #     self.ws.cell(row_n, col_n).value = value
    #     self.wb.save(self.filename)