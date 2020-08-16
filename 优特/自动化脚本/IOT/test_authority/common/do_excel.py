# -*- coding: utf-8 -*-
# @Time    : 2019/11/12 14:21
# @Author  : man.jiang
import os
import sys
import re
sys.path.append(os.path.dirname(os.path.abspath(__file__)) + "/../")
from common.read_config import get_common
from openpyxl import load_workbook
# from test03.test_case.test_case import TestHttp01

class do_excel:
    wb = load_workbook(get_common().get_excel()[0])#打开文件
    sheet = wb[get_common().get_excel()[1]]
    data = []
    max_row = sheet.max_row
    max_column = sheet.max_column
    #读取数据，转存为字典
    def get_date(self):
        for vars0 in range(2,self.max_row+1):
            item = {}
            item["http_url"] = "/auth-internet" + self.sheet.cell(vars0,2).value
            item["http_mthod"] = self.sheet.cell(vars0,3).value
            http_param  = self.sheet.cell(vars0,4).value
            http_param = http_param.replace('\n', '')
            http_param = http_param.replace('_x000D_', '')
            item["http_param"] = eval(http_param)
            item["case"] = self.sheet.cell(vars0,5).value
            item["code"] = self.sheet.cell(vars0,6).value
            self.data.append(item)
        return self.data
    #写入excel
    def into_date(self,row,column,value):
        self.sheet.cell(row = row, column = column, value = value)
        self.wb.save(get_common().get_excel()[0])#保存excel

if __name__ == '__main__':
    a = do_excel().get_date()
    print(a)
    # b = do_excel().write_data()
